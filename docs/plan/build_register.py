#!/usr/bin/env python3
"""Build the Etyme defect register workbook."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

FONT = "Arial"

INK      = "1F2A2B"
MUTED    = "5A6A6C"
HDR_BG   = "0F5C61"
HDR_FG   = "FFFFFF"
BAND     = "F1F5F5"
FATAL_BG = "F4DED9"; FATAL_FG = "8E2B1C"
SER_BG   = "F5EBD6"; SER_FG   = "7A5709"
FIX_BG   = "E3E9E9"; FIX_FG   = "44585A"
LIMIT_BG = "E6E2F1"; LIMIT_FG = "4A3B7A"
SOL_BG   = "DFEAE3"

thin = Side(style="thin", color="C8D2D2")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

# id, area, flow, severity, type, defect, consequence, solution, effort, when
ROWS = [
 ("D-01","Sign-up","Sign-up to first value","Fatal","Promise with no mechanism",
  "The homepage promises a scored bench within the hour. Behind the form there is no reply owner, no route for CVs to arrive, no parsing step and no requirement to score against.",
  "A vendor signs up, nothing happens, and they never come back. Every other loop in the product is downstream of this one.",
  "Keep the first ten vendors manual and make the manual fast. Add a 'paste a requirement you are working on now' field to the form. Return the scored shortlist as a link or PDF with no account required — the login comes after the value, never before.",
  "S","Before code"),

 ("D-02","Sign-up","Contacting consultants","Fatal","Missing flow + legal",
  "There is no consultant sign-up at all. They enter because a vendor uploaded a file about them, and the design then starts texting them. Nobody opted in.",
  "In the US, texting individuals without prior express consent carries statutory damages per message. Vendor assurance is not clearly a defence.",
  "First message is an opt-in, not a ping: it names the vendor, says why they are hearing from you, and STOP works from message one. Freshness pings start only after a yes. Store consent per person with timestamp and source. Vendor confirms the relationship at upload. Legal review before the first send.",
  "M","Before code"),

 ("D-03","Business","Agent share rising 30% to 88%","Fatal","Hope, not a mechanism",
  "Adding eval cases measures the problem but never fixes it. Nothing in the design changes a prompt, a tool or the routing, and no trigger, owner or schedule says who would.",
  "The margin curve carrying the whole business case never happens. Year four looks like year one at 45% instead of 84%.",
  "Make it a named weekly loop. Every Friday, one person takes the top failure cause by volume, makes one change, re-runs a frozen eval set, and keeps or reverts on first-pass yield. One hour, one person, written down.",
  "S","Before code"),

 ("D-04","Cross-cutting","Outcome feeding the learning loops","Fatal","Loop never closes",
  "Clients often never say why a candidate was rejected. The feedback arrow that every learning loop depends on arrives less than half the time.",
  "Triage never learns which requirements you win, vendor scores cannot be computed, and the rate benchmark starves.",
  "Treat silence as a label rather than missing data. Infer outcomes from observable signals: not viewed in 14 days, requirement closed with nobody hired, candidate placed elsewhere. Build the schema for inferred outcomes now — retrofitting it later is a rewrite.",
  "M","Before code"),

 ("D-05","Business","Five good submissions per day","Fatal","Cannot be diagnosed",
  "One number with no decomposition. Five different causes produce the same stalled figure and each needs an opposite fix.",
  "In month three, exactly when it matters, you cannot tell whether the problem is triage, the bench, consent, the check or requirement volume. You guess.",
  "Replace the single number with a six-stage funnel: requirements arrived, passed triage, had a bench match, consent given, passed the check, sent. The drop between stages names the cause.",
  "S","Before code"),

 ("D-06","Sign-up","The first scored bench","Serious","Missing precondition",
  "Scoring requires a live requirement. A brand-new vendor has none in the system, so the demonstration has no fuel.",
  "The moment that sells the product cannot be produced for a new customer.",
  "Collect a real requirement on the sign-up form. They have one open in another tab; it costs them ten seconds and it is exactly what the scoring needs.",
  "S","Before code"),

 ("D-07","Sign-up","An inbound lead arrives","Serious","Dead end",
  "A form submission has no owner, no reply-time target and no second attempt.",
  "Leads die quietly at 9pm on a Friday, and a dead lead teaches you nothing about who to approach next.",
  "One named person, a two-hour reply target during working hours, two follow-up attempts, then closed with a reason from a short list. The reason is the loop.",
  "S","Phase 1"),

 ("D-08","Business","Prime vendor and platform at once","Serious","Trust boundary",
  "You supply your own consultants while also running the platform. Standard master-vendor model, but sub-vendors will suspect front-running unless the rules are explicit.",
  "Sub-vendors never complain — they quietly stop sending their good people. Supply then caps at the size of your own bench, which is the thing sub-vendors exist to uncap.",
  "Label your own submissions in the client's inbox and put the arrangement in the contract. Then run the six supplier rules on the 'Supplier rules' tab: same clock, same score, same feedback, ranked on results, published share, categories you sit out.",
  "M","Phase 1"),

 ("D-09","Business","Free until it works","Serious","Undefined boundaries",
  "The trial has no defined start date and no defined end. Both sides will remember it differently.",
  "The disagreement surfaces at exactly the moment you are asking for money.",
  "Write it at sign-up: the clock starts on the first uploaded bench, ends at five good submissions a day or ninety days, whichever comes first, and both sides read the same dashboard throughout.",
  "S","Phase 1"),

 ("D-10","Vendor","Escalation after three failed checks","Serious","Dead end",
  "A person fixes the submission by hand and nothing about the cause is recorded.",
  "The identical failure returns next week, and the week after. The human becomes a permanent absorber of the same error.",
  "Every escalation closes with a cause chosen from a short list. Causes aggregate. When one reaches ten occurrences it becomes a work item against the agent — which is how the weekly improvement loop gets fed automatically.",
  "M","Phase 2"),

 ("D-11","Vendor","Consultant freshness","Serious","Collects rubbish",
  "Two pings with no reply marks someone unconfirmed, and then nothing. There is no archive rule and no path back to confirmed.",
  "Within a year a large share of the bench is people who left the market, still counted in the idle-bench-days figure the vendor dashboard leads with. The hero metric corrupts itself.",
  "Unconfirmed for 30 days raises one call task for a recruiter. Still nothing at 60 days, archive automatically and drop them from every count. One click restores them if they resurface.",
  "S","Phase 2"),

 ("D-12","Client","Requirement changes mid-flight","Serious","Missing state transition",
  "Nothing handles a rate moving, a location changing or a role being cancelled after submissions have been scored.",
  "Twenty scored submissions silently become wrong, and cancelled requirements keep burning compute downstream.",
  "Version the requirement. Stamp every score with the version it was scored against. On change, re-score what is in flight and flag stale shortlists. On cancel, hard-stop everything downstream immediately.",
  "M","Phase 2"),

 ("D-13","Client","Requirement quality check","Serious","Routed to the wrong person",
  "'Rate below market' is shown to whoever typed the requirement, usually a coordinator with no authority over pay.",
  "They cannot act, so they override. Within a month overriding is reflexive and the check is theatre.",
  "Route each problem to whoever can act on it — rate goes to the hiring manager with the market data attached. Track the override rate; above half, the check is failing rather than the user.",
  "M","Phase 2"),

 ("D-14","Client","Human sample review","Serious","No sampling discipline",
  "'Ten a week' with no rule for which ten. Left to a person they review the easy ones; left to 'most recent' the rare failures are never seen.",
  "The agreement rate looks excellent and means nothing, which is worse than having no check at all.",
  "Assign the sample, never let it be chosen. Stratify it: some random, some from the 70-80 score band where errors cluster, some shortlisted, and every case where the rules and the model disagreed.",
  "S","Phase 1"),

 ("D-15","Vendor","Consent — the NO branch","Fixable","Dead end",
  "The consent text asks yes or no. The design only says what happens on yes.",
  "A no leaves the submission in limbo and throws away the most valuable reply available — 'someone already put me forward there'.",
  "No cancels the submission, notifies the recruiter, and asks one follow-up: already submitted, rate too low, or not interested. That third answer is free deduplication and free market data.",
  "S","Phase 2"),

 ("D-16","Client","Deduplication","Fixable","Dead end",
  "A person judges 'not the same person' in the check queue. The verdict corrects that one record and vanishes.",
  "The rule that made the mistake is unchanged, so it makes the same false merge again.",
  "The human verdict becomes training data rather than a one-off correction. Feed disagreements back into the matching thresholds on a monthly review.",
  "M","Phase 3"),

 ("D-17","Client","Rate benchmark","Fixable","Circular reasoning",
  "Market rates are built from your own submissions, then used to tell clients their rate is below market.",
  "Early on your bench is the entire sample, so an expensive bench makes every client look underpriced.",
  "Do not use the benchmark below a minimum sample size per skill and city, and show the sample size beside every figure. Sanity-check the first year against published survey data.",
  "S","Phase 3"),

 ("D-18","Vendor","Submission check","Known limit","Check cannot verify truth",
  "The check confirms the CV asserts seven years of a skill. It cannot confirm the seven years happened. A well-written fabrication passes cleanly.",
  "If sold as fraud detection, the first fabricated CV that reaches a client ends the relationship.",
  "No fix exists and pretending otherwise is worse than admitting it. The check catches sloppiness; interviews catch lying. Never describe it as fraud detection in any sales conversation or on the site.",
  "-","Ongoing"),
]

FIRST = [
 (1,"Get the SMS consent question in front of a lawyer","D-02",
  "Before a single message goes out — not before launch, before the first message. Everything else here is a design decision you can revise. This one has statutory damages attached and no amount of good intent undoes a send.",
  "Written confirmation of the opt-in wording and the consent record you must keep."),
 (2,"Make the first hour real, by hand","D-01, D-06, D-07",
  "Add the requirement field to the sign-up form, name one person to reply within two hours, and send the scored shortlist back without asking anyone to create an account.",
  "Ten vendors onboarded manually, each with a scored shortlist delivered inside an hour."),
 (3,"Put a name and a Friday on the improvement loop","D-03",
  "One person, one hour a week, one change, measured against a frozen eval set, kept or reverted on the number.",
  "A dated log of changes with first-pass yield before and after each one."),
 (4,"Assume outcomes will not arrive, and infer them","D-04",
  "Treat silence as a label. Design the inferred-outcome signals into the schema now rather than retrofitting them.",
  "Outcome table accepts inferred as well as reported results, with the source recorded."),
 (5,"Build the funnel, not the number","D-05",
  "Six counters instead of one, on data you are already storing.",
  "The vendor dashboard shows the six-stage funnel, not a single submissions-per-day figure."),
]

RULES = [
 ("Same clock","Every supplier, including your own bench, is notified in the same second. No pre-release, ever.",
  "A thirty-minute head start is invisible to you and obvious to them within a month."),
 ("Same score","Your consultants run the identical check and scoring, with the same evidence shown.",
  "You already show the working on every match. That is your proof of fairness, and no other platform can offer it."),
 ("Same feedback","Every supplier sees why they lost and where they ranked, not only the winner.",
  "Without it they cannot improve, so they cannot win, so they leave. Same dead end as half the defects in this register."),
 ("Ranked on results","Distribution order comes from placement rate, never from ownership.",
  "The moment the order is 'us first', the vendor scorecard is decoration."),
 ("Published share","Show suppliers how often your own bench wins, per category, openly.",
  "They will work it out anyway. Publishing it is either a recruitment pitch or an early warning."),
 ("Categories you sit out","Name skill areas where your own bench does not bid at all.",
  "The cheapest trust you will ever buy, since you cannot cover every skill anyway."),
]

SEV_STYLE = {"Fatal":(FATAL_BG,FATAL_FG), "Serious":(SER_BG,SER_FG),
             "Fixable":(FIX_BG,FIX_FG), "Known limit":(LIMIT_BG,LIMIT_FG)}


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, size=9, color=HDR_FG)
        cell.fill = PatternFill("solid", fgColor=HDR_BG)
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = box
    ws.row_dimensions[row].height = 30


def build():
    wb = Workbook()

    # ---------------- Register ----------------
    ws = wb.active
    ws.title = "Register"

    ws["A1"] = "Etyme — defect register"
    ws["A1"].font = Font(name=FONT, bold=True, size=15, color=INK)
    ws["A2"] = ("Eighteen broken loops found by tracing every designed flow. "
                "Severity drives order. Fill in Owner, Status and Target date; leave the rest as the record of what was found.")
    ws["A2"].font = Font(name=FONT, size=9, color=MUTED)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:L2")
    ws.row_dimensions[2].height = 26

    ws["A3"] = "Legend — edit only the three yellow columns (J, K, L). Severity: Fatal changes what you build · Serious will hurt in month three · Fixable can wait · Known limit has no fix."
    ws["A3"].font = Font(name=FONT, size=8, italic=True, color=MUTED)
    ws.merge_cells("A3:L3")

    headers = ["ID","Area","Flow","Severity","Defect type","The defect",
               "Consequence if unfixed","Solution","Effort","Owner","Status","Target date"]
    hrow = 5
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    style_header(ws, hrow, len(headers))

    r = hrow + 1
    for i, row in enumerate(ROWS):
        (rid, area, flow, sev, dtype, defect, cons, sol, effort, when) = row
        vals = [rid, area, flow, sev, dtype, defect, cons, sol, effort, "", "Open", when]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = box
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
        # severity chip
        bg, fg = SEV_STYLE[sev]
        sc = ws.cell(row=r, column=4)
        sc.fill = PatternFill("solid", fgColor=bg)
        sc.font = Font(name=FONT, size=9, bold=True, color=fg)
        sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # id + solution emphasis
        ws.cell(row=r, column=1).font = Font(name=FONT, size=9, bold=True, color=INK)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="top")
        sol_cell = ws.cell(row=r, column=8)
        sol_cell.fill = PatternFill("solid", fgColor=SOL_BG)
        # editable columns
        for c in (10, 11, 12):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.row_dimensions[r].height = 92
        r += 1

    last = r - 1

    widths = {"A":8,"B":13,"C":26,"D":11,"E":22,"F":52,"G":46,"H":58,"I":8,"J":14,"K":12,"L":13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    dv_status = DataValidation(type="list", formula1='"Open,In progress,Done,Accepted risk"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add(f"K{hrow+1}:K{last}")

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A{hrow}:L{last}"

    # summary block
    s = last + 2
    ws.cell(row=s, column=1, value="Summary").font = Font(name=FONT, bold=True, size=11, color=INK)
    labels = ["Fatal","Serious","Fixable","Known limit","TOTAL","Open"]
    formulas = [
        f'=COUNTIF($D${hrow+1}:$D${last},"Fatal")',
        f'=COUNTIF($D${hrow+1}:$D${last},"Serious")',
        f'=COUNTIF($D${hrow+1}:$D${last},"Fixable")',
        f'=COUNTIF($D${hrow+1}:$D${last},"Known limit")',
        f'=COUNTA($A${hrow+1}:$A${last})',
        f'=COUNTIF($K${hrow+1}:$K${last},"Open")',
    ]
    for i, (lab, f) in enumerate(zip(labels, formulas)):
        rr = s + 1 + i
        lc = ws.cell(row=rr, column=1, value=lab)
        lc.font = Font(name=FONT, size=9, bold=(lab in ("TOTAL","Open")), color=INK)
        lc.border = box
        vc = ws.cell(row=rr, column=2, value=f)
        vc.font = Font(name=FONT, size=9, bold=(lab in ("TOTAL","Open")), color=INK)
        vc.alignment = Alignment(horizontal="center")
        vc.border = box
        if lab in SEV_STYLE:
            bg, fg = SEV_STYLE[lab]
            lc.fill = PatternFill("solid", fgColor=bg)
            lc.font = Font(name=FONT, size=9, bold=True, color=fg)
        ws.row_dimensions[rr].height = 15

    ws.cell(row=s, column=4,
            value="Counts are live formulas over the table above — they update as you change Status.")
    ws.cell(row=s, column=4).font = Font(name=FONT, size=8, italic=True, color=MUTED)

    ws.cell(row=hrow, column=6).comment = Comment(
        "Every entry here was found by tracing the designed flows, not by testing a running system. "
        "Nothing has been built yet, so these are design defects rather than observed bugs.", "Review", height=110, width=260)

    # ---------------- Fix first ----------------
    ws2 = wb.create_sheet("Fix first")
    ws2["A1"] = "Fix these five before any code"
    ws2["A1"].font = Font(name=FONT, bold=True, size=15, color=INK)
    ws2["A2"] = ("The other thirteen can be fixed as you go. These five change what gets built — "
                 "and the first two change what happens this week.")
    ws2["A2"].font = Font(name=FONT, size=9, color=MUTED)
    ws2.merge_cells("A2:E2")

    h2 = ["#","What to do","Defects it closes","Why it cannot wait","Done looks like"]
    for i, h in enumerate(h2, start=1):
        ws2.cell(row=4, column=i, value=h)
    style_header(ws2, 4, len(h2))

    rr = 5
    for i, (n, what, ids, why, done) in enumerate(FIRST):
        for c, v in enumerate([n, what, ids, why, done], start=1):
            cell = ws2.cell(row=rr, column=c, value=v)
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = box
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
        ws2.cell(row=rr, column=1).font = Font(name=FONT, size=13, bold=True, color=FATAL_FG)
        ws2.cell(row=rr, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws2.cell(row=rr, column=2).font = Font(name=FONT, size=10, bold=True, color=INK)
        ws2.cell(row=rr, column=5).fill = PatternFill("solid", fgColor=SOL_BG)
        ws2.row_dimensions[rr].height = 78
        rr += 1

    for col, w in {"A":5,"B":38,"C":18,"D":58,"E":46}.items():
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A5"

    # ---------------- Supplier rules ----------------
    ws3 = wb.create_sheet("Supplier rules")
    ws3["A1"] = "Running as prime vendor without losing your sub-vendors"
    ws3["A1"].font = Font(name=FONT, bold=True, size=15, color=INK)
    ws3["A2"] = ("Policy, not defects — these close D-08. Supplying your own consultants while running the platform is the "
                 "master vendor model and is entirely normal. The risk is not the client objecting, it is sub-vendors quietly leaving.")
    ws3["A2"].font = Font(name=FONT, size=9, color=MUTED)
    ws3["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws3.merge_cells("A2:C2")
    ws3.row_dimensions[2].height = 28

    h3 = ["Rule","What it means in the product","Why they will not stay without it"]
    for i, h in enumerate(h3, start=1):
        ws3.cell(row=4, column=i, value=h)
    style_header(ws3, 4, len(h3))

    rr = 5
    for i, (rule, means, why) in enumerate(RULES):
        for c, v in enumerate([rule, means, why], start=1):
            cell = ws3.cell(row=rr, column=c, value=v)
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = box
            if i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=BAND)
        ws3.cell(row=rr, column=1).font = Font(name=FONT, size=10, bold=True, color=INK)
        ws3.row_dimensions[rr].height = 56
        rr += 1

    note = rr + 1
    ws3.cell(row=note, column=1, value="The test of whether you mean it")
    ws3.cell(row=note, column=1).font = Font(name=FONT, bold=True, size=10, color=INK)
    ws3.cell(row=note+1, column=1,
             value=("On your own consultant you bill $85, pay $60 and keep $25. On a sub-vendor placement you might keep $10 — "
                    "but with no bench to carry, no idle days and no ceiling on coverage. The blended business beats either alone, "
                    "which means you should genuinely want sub-vendors to win sometimes. If that sentence sits badly, the model will not hold, "
                    "because every rule above costs you a placement you could have taken."))
    ws3.cell(row=note+1, column=1).font = Font(name=FONT, size=9, color=INK)
    ws3.cell(row=note+1, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws3.merge_cells(start_row=note+1, start_column=1, end_row=note+1, end_column=3)
    ws3.row_dimensions[note+1].height = 58
    ws3.cell(row=note+1, column=1).fill = PatternFill("solid", fgColor=SOL_BG)

    for col, w in {"A":22,"B":58,"C":58}.items():
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A5"

    out = "/tmp/claude-0/-home-user-etyme-2017/594d6bc2-b39b-575d-abf9-941b742ac6f3/scratchpad/Etyme-Defect-Register.xlsx"
    wb.save(out)
    print("saved:", out)
    print("register rows:", len(ROWS))


if __name__ == "__main__":
    build()
